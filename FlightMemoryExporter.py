from bs4 import BeautifulSoup
import openpyxl

from selenium import webdriver
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException


import tkinter
from tkinter import ttk
from tkinter import filedialog
import threading
import chromedriver_autoinstaller


#Button press ####################################################################################################

def OK():
    saveDir = filedialog.askdirectory()
    account = eUsername.get()
    password = ePassword.get()
    mainThread = threading.Thread(target=run, args=(account, password, saveDir))
    mainThread.run()

#Run###############################################################################################################
def run(account, password, saveDir):
    print(f"Directory chosen:{saveDir}")
    #Driver Setup
    chromedriver_autoinstaller.install()
    print('chromedriver installed')
    driver = webdriver.Chrome()
    #OpenLoginPage
    LOGIN_PAGE = "https://www.flightmemory.com/"

    driver.get(LOGIN_PAGE)
    wait = WebDriverWait(driver, 5)
    #EnterName
    wait.until(expected_conditions.element_to_be_clickable((By.NAME, "username"))).send_keys(account)
    print('enteredName')
    #EnterPassword
    wait.until(expected_conditions.element_to_be_clickable((By.NAME, "passwort"))).send_keys(password)
    print('enteredPassword')
    #SignIn
    wait.until(expected_conditions.element_to_be_clickable((By.XPATH, ".//input[@value='SignIn' and @type='submit']"))).click()
    print("SignedIn")
    #UpdateProgressBar
    pb['value'] = 40
    root.update_idletasks()

    #SeeIfPageSignedInSuccesfully
    try:
        wait.until(expected_conditions.element_to_be_clickable((By.XPATH, "//*[contains(text(), 'FLIGHTDATA')]"))).click()
    except TimeoutException:
        popupmsg('Wrong username or password')
        return

    print('found flightdata')
    #CreatePages array
    pages = []
    pages.append(driver.execute_script("return document.documentElement.outerHTML"))
    #UpdateProgressBar
    pb['value'] = 60
    root.update_idletasks()

    #GetNumber of pages
    while True:
        try:
            WebDriverWait(driver, 10).until(
                expected_conditions.element_to_be_clickable((By.XPATH, "//img[contains(@src,'/images/next.gif')]"))).click()
            pages.append(driver.execute_script("return document.documentElement.outerHTML"))
        except TimeoutException:
            print(f"Pages Found:{int(len(pages))}")
            break

    #update progressbar
    pb['value'] = 90
    root.update_idletasks()

    # Create workbook object
    wb = openpyxl.Workbook()
    sheet = wb.sheetnames[0]
    sheet = wb[sheet]
    sheet.title = 'Flights'

    # table ifo as in myflightradar
    sheet.cell(row=1, column=1).value = 'Date'
    sheet.cell(row=1, column=2).value = 'Flight number'
    sheet.cell(row=1, column=3).value = 'From'
    sheet.cell(row=1, column=4).value = 'To'
    sheet.cell(row=1, column=5).value = 'Dep time'
    sheet.cell(row=1, column=6).value = 'Arr time'
    sheet.cell(row=1, column=7).value = 'Duration'
    sheet.cell(row=1, column=8).value = 'Airline'
    sheet.cell(row=1, column=9).value = 'Aircraft'
    sheet.cell(row=1, column=10).value = 'Registration'
    sheet.cell(row=1, column=11).value = 'Seat number'
    sheet.cell(row=1, column=12).value = 'Seat type'
    sheet.cell(row=1, column=13).value = 'Flight class'
    sheet.cell(row=1, column=14).value = 'Flight reason'
    sheet.cell(row=1, column=15).value = 'Note'
    sheet.cell(row=1, column=16).value = 'Dep_id'
    sheet.cell(row=1, column=17).value = 'Arr_id'
    sheet.cell(row=1, column=18).value = 'Airline_id'
    sheet.cell(row=1, column=19).value = 'Aircraft_id'

    global lastidx
    lastidx = 0

    for page in pages:
        # find table with BS
        data = BeautifulSoup(page, 'html.parser')
        container = data.select_one('.container')
        tablebody = container.find_all('tbody')[2]
        trs = tablebody.find_all('tr', recursive=False)
        print(page)

        # loop through each line
        for idx, tr in enumerate(trs):
            idx += lastidx
            idx += 1

            if idx != lastidx + 1:
                # get info
                flight_number = tr.find_all('td')[0]
                date_dep_arr = tr.find_all('td')[1]
                Departure = tr.find_all('td')[2]
                Arrival = tr.find_all('td')[4]
                Distance = tr.find_all('td')[6]
                FlightTime = tr.find_all('td')[8]
                Airline_Flightinfo = tr.find_all('td')[10]
                Airplane = tr.find_all('td')[11]
                Seat = tr.find_all('td')[12]
                # edit time formating
                # '-'.join([date_dep_arr[i:i+10] for i in range(0, len(date_dep_arr), 3)])
                # addInfoToSheetMyFlightRadar
                date = date_dep_arr.get_text()[0:10]
                date = date.replace('.', '/')
                sheet.cell(row=idx, column=1).value = date
                sheet.cell(row=idx, column=2).value = getinfo(Airline_Flightinfo, 1)
                sheet.cell(row=idx, column=3).value = Departure.get_text()
                sheet.cell(row=idx, column=4).value = Arrival.get_text()
                sheet.cell(row=idx, column=5).value = date_dep_arr.get_text()[10:15]
                sheet.cell(row=idx, column=6).value = date_dep_arr.get_text()[15:20]
                sheet.cell(row=idx, column=7).value = FlightTime.get_text()
                sheet.cell(row=idx, column=8).value = getinfo(Airline_Flightinfo, 0)
                sheet.cell(row=idx, column=9).value = getinfo(Airplane, 0)
                sheet.cell(row=idx, column=10).value = getinfo(Airplane, 1)
                sheet.cell(row=idx, column=11).value = Seat.get_text().split('/')[0]
                sheet.cell(row=idx, column=12).value = getSeatInfo(1, Seat)
                sheet.cell(row=idx, column=13).value = getSeatInfo(2, Seat)
                sheet.cell(row=idx, column=14).value = getSeatInfo(3, Seat)

            print(idx)
        lastidx = idx

    # remove empty rows
    for row in range(sheet.max_row + 1, 1, -1):  # range is from bottom to top, step -1
        if sheet[row][1].value is None:
            sheet.delete_rows(idx=row, amount=1)
    # Save file + popup
    wb.save((saveDir + '\Flights.xlsx'))
    pb['value'] = 100
    root.update_idletasks()
    popupmsg('finished')
    print('finished')
#GeiInfo separated by br##########################################################################################
def getinfo(data, index):
    datatextarray = data.get_text(separator='|', strip=True).split('|')
    x = len(datatextarray) - 1
    if index > x:
        datatextarray.append('')
        text = datatextarray[index]
        return text
    else:
        text = datatextarray[index]
        return text
#Get seat info #########################################################################################################
def getSeatInfo(InfoType, data):
    data = data.get_text()
    if InfoType == 1:
        if 'Window' in data:
            return ('Window')  # 1
        elif 'Middle' in data:
            return ('Middle')  # 2
        elif 'Aisle' in data:
            return ('Aisle')  # 3
        else:
            return (' ')  # 0
    elif InfoType == 2:
        if 'EconomyPlus' in data:
            return ('Economy Plus')  # 4
        elif 'Economy' in data:
            return ('Economy')  # 1
        elif 'Business' in data:
            return ('Business')  # 2
        elif 'First' in data:
            return ('First')  # 3

        else:
            return ('0')
    elif InfoType == 3:
        if 'Personal' in data:
            return ('Personal')  # 1
        else:
            # need to learn other type numbers
            return (' ')  # 0
#PopUp############################################################################################################
def popupmsg(msg):
    popup = tkinter.Tk()
    popup.wm_title("!")
    label = ttk.Label(popup, text=msg)
    label.pack(side="top", fill="x", pady=10)
    B1 = ttk.Button(popup, text="Okay", command = popup.destroy)
    B1.pack()
    popup.mainloop()

#GUI###############################################################################################################
root = tkinter.Tk()

#UsernameInput
tkinter.Label(root, text='Username:').grid(row = 0, column = 0)
eUsername = tkinter.Entry(root, width=15, borderwidth=2)
eUsername.grid(row = 0, column = 1)
#PasswordInput
tkinter.Label(root, text='Password:').grid(row = 1, column=0)
ePassword = tkinter.Entry(root, width=15, borderwidth=2, show='*')
ePassword.grid(row = 1, column = 1)
#progressbar
pb = ttk.Progressbar(orient='horizontal', length=100, mode='determinate')
pb.grid(row = 3, column = 1)
directory = './'
#ConfirmButton
tkinter.Button(root, text="OK",  command=lambda : OK()).grid(row=3, column=0)

#####################################################################################################################







root.mainloop()